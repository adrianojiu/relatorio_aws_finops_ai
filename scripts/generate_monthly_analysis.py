#!/usr/bin/env python3
"""
Gera análise mensal consolidada de custos AWS a partir de:
  - CSVs mensais existentes em export_monthly/
  - Payloads Bedrock diários em output/YYYY-MM-DD/*_bedrock_payload.json
  - Calendário de eventos em prompts/assets/Régua de Pushs_SMS Now Online.xlsx

Saída: output/monthly/relatorio_mensal_YYYY-MM.txt

Uso (EC2 com IAM role):
  python3 scripts/generate_monthly_analysis.py --month 2026-04 --enable-bedrock

Uso (local com perfil AWS CLI):
  python3 scripts/generate_monthly_analysis.py --month 2026-04 \
    --aws-profile prd-ciam --enable-bedrock
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import defaultdict
from datetime import datetime, date, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import boto3
from botocore.config import Config as BotoConfig

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
for _p in (str(PROJECT_ROOT), str(SRC_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from src import config  # noqa: E402  (namespace package — src/ sem __init__.py)
from integrations.bedrock import (  # noqa: E402
    _build_request_payload,
    _extract_response_text,
)


# ──────────────────────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera análise mensal de custos AWS com correlação de eventos e IA."
    )
    parser.add_argument(
        "--month", required=True,
        help="Mês no formato YYYY-MM. Ex.: 2026-04",
    )
    parser.add_argument("--aws-profile", default=config.AWS_PROFILE)
    parser.add_argument("--bedrock-region", default=config.BEDROCK_REGION)
    parser.add_argument("--bedrock-model", default=config.BEDROCK_MODEL_ID)
    parser.add_argument(
        "--enable-bedrock", action="store_true", default=False,
        help="Habilita análise via Bedrock (gera narrativa executiva do mês).",
    )
    parser.add_argument(
        "--output-dir", default=str(PROJECT_ROOT / "output" / "monthly"),
        help="Diretório de saída. Padrão: <project_root>/output/monthly",
    )
    parser.add_argument(
        "--export-dir", default=config.MONTHLY_EXPORT_OUTPUT_DIR,
        help=f"Diretório dos CSVs mensais. Padrão: {config.MONTHLY_EXPORT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--sns-topic-arn",
        default=config.SNS_TOPIC_ARN,
        help="ARN do tópico SNS para envio do relatório por e-mail. Padrão: env FINOPS_SNS_TOPIC_ARN",
    )
    parser.add_argument(
        "--s3-bucket",
        default=config.S3_REPORT_BUCKET,
        help="Bucket S3 para upload dos arquivos antes do envio SNS. Padrão: env FINOPS_S3_BUCKET",
    )
    parser.add_argument(
        "--s3-prefix",
        default=config.S3_REPORT_PREFIX_MONTHLY,
        help=f"Prefixo S3 para organização dos arquivos. Padrão: {config.S3_REPORT_PREFIX_MONTHLY}",
    )
    return parser.parse_args()


# ──────────────────────────────────────────────────────────────────────────────
# CSV loading
# ──────────────────────────────────────────────────────────────────────────────

def _resolve_csv_path(export_dir: Path, filename_template: str, month: str) -> Path:
    stem = Path(filename_template).stem
    suffix = Path(filename_template).suffix
    return export_dir / f"{stem}-{month}{suffix}"


def load_monthly_csv(path: Path) -> Tuple[Dict[str, Decimal], Dict[str, Dict[str, Decimal]]]:
    """Lê um CSV mensal e retorna (totais_por_servico, {dia: {servico: valor}}).

    Ignora a coluna 'Total costs' do CSV — calculamos internamente.
    """
    service_totals: Dict[str, Decimal] = {}
    by_day: Dict[str, Dict[str, Decimal]] = {}

    if not path.exists():
        return service_totals, by_day

    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        raw_header = next(reader)

    # Mapeia índice → nome do serviço, pulando a coluna "Total costs"
    service_cols: List[Tuple[int, str]] = []
    for idx, col in enumerate(raw_header[1:], start=1):
        name = col.replace("($)", "").strip()
        if name and name != "Total costs":
            service_cols.append((idx, name))

    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        next(reader)  # pula header
        for row in reader:
            if not row:
                continue
            key = row[0].strip()
            if key == "Service total":
                for col_idx, svc in service_cols:
                    val = row[col_idx].strip() if col_idx < len(row) else ""
                    if val:
                        service_totals[svc] = Decimal(val)
            elif key.startswith("20"):  # linha de data
                day_data: Dict[str, Decimal] = {}
                for col_idx, svc in service_cols:
                    val = row[col_idx].strip() if col_idx < len(row) else ""
                    if val:
                        try:
                            day_data[svc] = Decimal(val)
                        except Exception:
                            pass
                if day_data:
                    by_day[key] = day_data

    return service_totals, by_day


# ──────────────────────────────────────────────────────────────────────────────
# Daily payload loading
# ──────────────────────────────────────────────────────────────────────────────

def load_daily_payloads(month: str, output_root: Path) -> Dict[str, List[dict]]:
    """Varre output/YYYY-MM-DD/ para o mês e extrai o resumo de anomalias de cada dia."""
    result: Dict[str, List[dict]] = {}
    year, mon = month.split("-")

    for day_dir in sorted(output_root.glob(f"{year}-{mon}-*")):
        if not day_dir.is_dir():
            continue
        payload_files = list(day_dir.glob("*_bedrock_payload.json"))
        if not payload_files:
            continue
        try:
            data = json.loads(payload_files[0].read_text(encoding="utf-8"))
        except Exception:
            continue

        # Usa reference_day do payload; fallback: nome da pasta
        ref_date = (
            data.get("period", {}).get("reference_day")
            or data.get("period", {}).get("anchor_day")
            or day_dir.name
        )

        anomalies = [
            {
                "service": a.get("service"),
                "usage_type": a.get("usage_type"),
                "cost_reference_day": a.get("cost_reference_day"),
                "avg_7d": a.get("avg_7d"),
                "delta_pct": a.get("delta_pct"),
                "delta_usd": a.get("delta_usd"),
                "hypothesis": a.get("hypothesis"),
            }
            for a in data.get("anomalies", [])
        ]
        result[ref_date] = anomalies

    return result


# ──────────────────────────────────────────────────────────────────────────────
# Event calendar loading
# ──────────────────────────────────────────────────────────────────────────────

_MONTH_NAME_MAP = {
    "01": "JANEIRO", "02": "FEVEREIRO", "03": "MARÇO", "04": "ABRIL",
    "05": "MAIO", "06": "JUNHO", "07": "JULHO", "08": "AGOSTO",
    "09": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO",
}


def load_events_for_month(month: str) -> Dict[str, List[dict]]:
    """Lê a planilha de eventos e retorna {data: [{time, alcance, nodes, description}]}."""
    events: Dict[str, List[dict]] = defaultdict(list)

    if not HAS_OPENPYXL:
        print("[monthly] AVISO: openpyxl não instalado — calendário de eventos ignorado.")
        return dict(events)

    cal_path = Path(config.BUSINESS_EVENT_CALENDAR_FILE)
    if not cal_path.exists():
        print(f"[monthly] AVISO: planilha não encontrada em {cal_path}")
        return dict(events)

    try:
        wb = openpyxl.load_workbook(str(cal_path), data_only=True)
    except Exception as exc:
        print(f"[monthly] AVISO: erro ao abrir planilha — {exc}")
        return dict(events)

    year, mon = month.split("-")
    sheet_name = f"{_MONTH_NAME_MAP.get(mon, '')} {year}"
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else (wb["ATUAL"] if "ATUAL" in wb.sheetnames else wb.active)

    start_date = date(int(year), int(mon), 1)
    next_month_start = (start_date.replace(day=28) + timedelta(days=4)).replace(day=1)

    for row in sheet.iter_rows(values_only=True):
        data_col = row[2] if len(row) > 2 else None
        if not isinstance(data_col, datetime):
            continue
        d = data_col.date()
        if not (start_date <= d < next_month_start):
            continue

        horario = row[3] if len(row) > 3 else None
        alcance = row[5] if len(row) > 5 else None
        nodes = row[6] if len(row) > 6 else None
        destaque = row[7] if len(row) > 7 else None

        events[d.isoformat()].append({
            "time": str(horario) if horario else "",
            "alcance": str(alcance).strip() if alcance else "",
            "nodes": int(nodes) if isinstance(nodes, (int, float)) else 0,
            "description": str(destaque)[:120].strip() if destaque else "",
        })

    return dict(events)


# ──────────────────────────────────────────────────────────────────────────────
# Monthly summary building
# ──────────────────────────────────────────────────────────────────────────────

def _ranked_services(service_totals: Dict[str, Decimal], top_n: int = 12) -> List[dict]:
    grand_total = sum(service_totals.values()) or Decimal("1")
    ranked = sorted(service_totals.items(), key=lambda x: -x[1])
    return [
        {
            "service": svc,
            "total_usd": float(round(amt, 2)),
            "pct": float(round(amt / grand_total * 100, 1)),
        }
        for svc, amt in ranked[:top_n]
        if amt > 0
    ]


def _recurring_anomalies(anomalies_by_day: Dict[str, List[dict]]) -> List[dict]:
    """Identifica serviços que aparecem como anomalia em mais de um dia de referência."""
    counter: Dict[str, dict] = defaultdict(lambda: {"count": 0, "total_delta_usd": 0.0, "days": []})
    for day, anomalies in anomalies_by_day.items():
        for a in anomalies:
            key = f"{a['service']} | {a['usage_type'] or '-'}"
            counter[key]["count"] += 1
            counter[key]["total_delta_usd"] += float(a.get("delta_usd") or 0)
            counter[key]["days"].append(day)
    return sorted(
        [{"key": k, **v} for k, v in counter.items() if v["count"] > 1],
        key=lambda x: -x["count"],
    )


def build_monthly_summary(
    month: str,
    all_totals: Dict[str, Decimal],
    all_by_day: Dict[str, Dict[str, Decimal]],
    pdp_totals: Dict[str, Decimal],
    pdp_by_day: Dict[str, Dict[str, Decimal]],
    anomalies_by_day: Dict[str, List[dict]],
    events_by_day: Dict[str, List[dict]],
) -> dict:
    grand_total = float(round(sum(all_totals.values()), 2))
    pdp_total = float(round(sum(pdp_totals.values()), 2))

    # Tax é consolidada pela AWS no primeiro dia do mês — isolá-la por dia
    # evita que esse dia apareça como outlier na leitura da curva diária.
    tax_by_day = {
        day: float(round(svcs.get("Tax", Decimal("0")), 2))
        for day, svcs in all_by_day.items()
    }

    # Custo total por dia (todos os serviços, incluindo Tax)
    cost_by_day = {
        day: float(round(sum(svcs.values()), 2))
        for day, svcs in sorted(all_by_day.items())
    }
    # Custo sem Tax — usado como referência de curva para não distorcer o dia com Tax consolidada
    cost_ex_tax_by_day = {
        day: float(round(sum(svcs.values()) - svcs.get("Tax", Decimal("0")), 2))
        for day, svcs in sorted(all_by_day.items())
    }
    pdp_cost_by_day = {
        day: float(round(sum(svcs.values()), 2))
        for day, svcs in sorted(pdp_by_day.items())
    }

    # Detalhe por dia com contexto de eventos
    all_days = sorted(set(cost_by_day) | set(events_by_day))
    days_detail = []
    for d in all_days:
        evts = events_by_day.get(d, [])
        max_nodes = max((e["nodes"] for e in evts), default=0)
        ggg_count = sum(1 for e in evts if e.get("alcance") == "GGG")
        tax = tax_by_day.get(d, 0.0)
        days_detail.append({
            "date": d,
            "total_cost": cost_by_day.get(d, 0.0),
            "cost_ex_tax": cost_ex_tax_by_day.get(d, 0.0),
            "tax": tax,
            "pdp_cost": pdp_cost_by_day.get(d, 0.0),
            "events_count": len(evts),
            "max_nodes": max_nodes,
            "ggg_count": ggg_count,
            "has_payload": d in anomalies_by_day,
        })

    days_without_payload = [
        d["date"] for d in days_detail
        if not d["has_payload"] and d["total_cost"] > 0
    ]

    return {
        "month": month,
        "total_cost_usd": grand_total,
        "pdp_cost_usd": pdp_total,
        "pdp_pct_of_total": round(pdp_total / grand_total * 100, 1) if grand_total else 0.0,
        "days_in_month": len(cost_by_day),
        "avg_daily_cost_usd": round(grand_total / len(cost_by_day), 2) if cost_by_day else 0.0,
        "top_services_all": _ranked_services(all_totals, 12),
        "top_services_pdp": _ranked_services(pdp_totals, 8),
        "cost_by_day": cost_by_day,
        "pdp_cost_by_day": pdp_cost_by_day,
        "days_detail": days_detail,
        "days_without_payload": days_without_payload,
        "anomalies_by_day": anomalies_by_day,
        "recurring_anomalies": _recurring_anomalies(anomalies_by_day),
        "events_by_day": events_by_day,
    }


# ──────────────────────────────────────────────────────────────────────────────
# Bedrock
# ──────────────────────────────────────────────────────────────────────────────

def _load_bedrock_context() -> str:
    """Extrai a seção 'Contexto para Bedrock' do PROJECT_CONTEXT.md."""
    ctx_path = Path(config.PROJECT_CONTEXT_FILE)
    if not ctx_path.exists():
        return ""
    text = ctx_path.read_text(encoding="utf-8")
    marker = "## Contexto para Bedrock"
    idx = text.find(marker)
    if idx < 0:
        return ""
    end_idx = text.find("\n## ", idx + len(marker))
    return text[idx:end_idx].strip() if end_idx > 0 else text[idx:].strip()


def _build_monthly_prompt(summary: dict, context_operational: str) -> str:
    return (
        "Você é um analista sênior de FinOps AWS. Analise o relatório mensal estruturado abaixo "
        "e produza uma análise executiva consolidada do mês.\n\n"
        "A análise deve conter obrigatoriamente:\n"
        "1. **Resumo executivo**: custo total, participação do PDP, média diária e tendência geral.\n"
        "2. **Top drivers de custo**: os 3-5 serviços que mais pesaram e por quê (use top_services_all).\n"
        "3. **Correlação custo × eventos Claro TV+**: como os eventos (GGG/GG/G) influenciaram os dias de pico "
        "— compare custo médio nos dias com GGG vs. dias sem GGG.\n"
        "4. **Anomalias recorrentes**: serviços em recurring_anomalies — indício de tendência estrutural vs. pontual.\n"
        "5. **Semanas de alta e baixa**: identifique 2-3 períodos distintos e suas causas.\n"
        "6. **Pontos de atenção para o próximo mês**: o que monitorar com base no padrão observado.\n\n"
        "Regras de resposta:\n"
        "- Resposta em português, formato markdown com seções claras.\n"
        "- Objetivo e executivo — não repita dados brutos que já estão no relatório.\n"
        "- Diferencie desvio esperado (evento programado) de anomalia real (sem explicação de negócio).\n"
        "- Não force causalidade sem evidência no payload.\n"
        "- Se o mês analisado for 2026-04, mencione o stress test de 15/04 (00h–05h UTC-3) ao interpretar "
        "o custo desse dia.\n\n"
        "Regras de domínio obrigatórias (não viole nenhuma delas):\n"
        "- PDP significa Policy Decision Point — componente de autorização do Ping Identity em EKS. "
        "NUNCA expanda PDP como 'Plataforma de Distribuição de Push' ou qualquer outra variação; isso está incorreto.\n"
        "- AWS End User Messaging é usado exclusivamente para mensagens transacionais: OTP, MFA, confirmações "
        "de cadastro e reset de senha. Não é usado para campanhas, notificações push de broadcast nem CRM.\n"
        "- NUNCA associe variação de End User Messaging/SMS a eventos do Claro TV+, audiência ou notificações "
        "push de broadcast. SMS é acionado por ações do usuário, não por programação de eventos.\n"
        "- Evidência histórica (abril/2026): anomalias de SMS tiveram delta acumulado negativo (-$119,60), "
        "ocorreram em dias sem GGG e estiveram ausentes em vários dias GGG — a volatilidade é transacional.\n"
        "- ELB/LCU sobe como efeito derivado do volume de requisições (EC2/EKS); quando ELB e EC2 subirem "
        "juntos em dia de evento, classifique como efeito em cascata com driver único de tráfego, "
        "não como anomalias independentes.\n"
        "- Transit Gateway é sinal fraco para correlacionar com eventos do Claro TV+ ou scaling de EKS; "
        "não use como driver principal quando ELB e EC2 já explicam o aumento.\n\n"
        f"Contexto operacional do ambiente:\n{context_operational}\n\n"
        f"Dados estruturados do mês:\n{json.dumps(summary, indent=2, ensure_ascii=False)}"
    )


def invoke_bedrock(
    prompt: str,
    model_id: str,
    aws_profile: str | None,
    bedrock_region: str,
) -> Optional[str]:
    """Invoca o Bedrock e retorna o texto da resposta."""
    # aws_profile=None usa credenciais do ambiente (role EC2, env vars, etc.)
    session = boto3.Session(profile_name=aws_profile or None, region_name=bedrock_region)
    client = session.client(
        "bedrock-runtime",
        config=BotoConfig(connect_timeout=60, read_timeout=600),
    )
    payload = _build_request_payload(model_id, prompt)
    response = client.invoke_model(
        modelId=model_id,
        contentType="application/json",
        accept="application/json",
        body=json.dumps(payload),
    )
    body = json.loads(response["body"].read())
    return _extract_response_text(model_id, body)


# ──────────────────────────────────────────────────────────────────────────────
# TXT report
# ──────────────────────────────────────────────────────────────────────────────

def _bar(value: float, max_value: float, width: int = 28) -> str:
    if max_value <= 0:
        return ""
    filled = int(round(value / max_value * width))
    return "█" * filled + "░" * (width - filled)


def write_txt_report(
    path: Path,
    summary: dict,
    ai_text: Optional[str],
    model_id: str,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    month = summary["month"]
    lines: List[str] = []

    # ── Cabeçalho ──────────────────────────────────────────────────────────────
    lines += [
        "=" * 72,
        f"  RELATÓRIO MENSAL DE CUSTOS AWS — {month}",
        "=" * 72,
        "",
        f"  Custo total do mês  : $ {summary['total_cost_usd']:>12,.2f}",
        f"  Custo PDP           : $ {summary['pdp_cost_usd']:>12,.2f}  ({summary['pdp_pct_of_total']:.1f}% do total — já incluso no total acima)",
        f"  Média diária        : $ {summary['avg_daily_cost_usd']:>12,.2f}",
        f"  Dias com dados CSV  : {summary['days_in_month']}",
        f"  Dias com payload    : {summary['days_in_month'] - len(summary['days_without_payload'])}",
        "",
    ]

    # ── Top serviços — todos ────────────────────────────────────────────────────
    days_in_month = summary["days_in_month"] or 1
    grand_total = summary["total_cost_usd"]
    col_header = f"  {'Serviço':<42} {'Mês($)':>10}  {'Média/dia':>9}  {'%total':>6}"
    lines += [
        "─" * 72,
        "  TOP SERVIÇOS DO MÊS  — cada linha é uma parcela do custo total",
        "  A soma de todos os serviços = custo total do mês ($)",
        col_header,
        "─" * 72,
    ]
    max_svc = summary["top_services_all"][0]["total_usd"] if summary["top_services_all"] else 1
    shown_total = 0.0
    for s in summary["top_services_all"]:
        bar = _bar(s["total_usd"], max_svc)
        avg_day = s["total_usd"] / days_in_month
        shown_total += s["total_usd"]
        lines.append(
            f"  {s['service']:<42} ${s['total_usd']:>10,.2f}  ${avg_day:>7,.2f}  {s['pct']:>5.1f}%  {bar}"
        )
    remaining = grand_total - shown_total
    if remaining > 0.01:
        avg_rem = remaining / days_in_month
        pct_rem = remaining / grand_total * 100
        lines.append(f"  {'(demais serviços)':<42} ${remaining:>10,.2f}  ${avg_rem:>7,.2f}  {pct_rem:>5.1f}%")
    lines += [
        f"  {'─'*42}   {'─'*10}",
        f"  {'TOTAL DO MÊS':<42} ${grand_total:>10,.2f}  ${grand_total/days_in_month:>7,.2f}  {'100.0%':>6}",
        "",
    ]

    # ── Top serviços — PDP ─────────────────────────────────────────────────────
    if summary["top_services_pdp"]:
        pdp_total = summary["pdp_cost_usd"]
        lines += [
            "─" * 72,
            "  TOP SERVIÇOS PDP  — parcelas do custo PDP já contidas no Total do mês",
            "  %total = % em relação ao custo PDP; os valores já estão incluídos no custo total acima",
            col_header,
            "─" * 72,
        ]
        max_pdp = summary["top_services_pdp"][0]["total_usd"] if summary["top_services_pdp"] else 1
        for s in summary["top_services_pdp"]:
            bar = _bar(s["total_usd"], max_pdp)
            avg_day = s["total_usd"] / days_in_month
            lines.append(
                f"  {s['service']:<42} ${s['total_usd']:>10,.2f}  ${avg_day:>7,.2f}  {s['pct']:>5.1f}%  {bar}"
            )
        lines += [
            f"  {'─'*42}   {'─'*10}",
            f"  {'TOTAL PDP':<42} ${pdp_total:>10,.2f}  ${pdp_total/days_in_month:>7,.2f}",
            "",
        ]

    # ── Custo diário × eventos ─────────────────────────────────────────────────
    lines += [
        "─" * 72,
        "  CUSTO DIÁRIO × EVENTOS  (◆ = dia com evento GGG  · = sem payload)",
        "  Total($) = soma de todos os serviços no dia (inclui PDP)",
        "  PDP($)   = parcela do Total atribuída ao PDP — já está contida no Total, não é somada",
        "  Curva comparativa usa custo sem Tax — Tax consolidada pela AWS no 1º dia distorceria a leitura",
        "  † = dia com Tax consolidada; o custo real do dia está entre parênteses",
        "─" * 72,
        f"  {'Data':<12} {'s/Tax($)':>10}  {'PDP($)':>9}  {'Nodes':>5}  Eventos",
        "",
    ]
    # Curva de referência sem Tax para não distorcer dias com Tax consolidada
    curve_values = [d["cost_ex_tax"] for d in summary["days_detail"] if d["cost_ex_tax"] > 0]
    max_curve = max(curve_values) if curve_values else 1
    avg_curve = sum(curve_values) / len(curve_values) if curve_values else 0

    for d in summary["days_detail"]:
        evts = summary["events_by_day"].get(d["date"], [])
        ggg_marker = "◆" if d["ggg_count"] > 0 else " "
        no_payload = " ·" if not d["has_payload"] else "  "
        evt_summary = ", ".join(
            f"{e['alcance']}({e['nodes']}n)" for e in evts[:3]
        ) if evts else "—"

        tax = d.get("tax", 0.0)
        cost_ex_tax = d["cost_ex_tax"]

        if tax > 0.01:
            # Dia com Tax: mostra custo sem Tax para curva, total real entre parênteses
            cost_str = f"${cost_ex_tax:>9,.2f}†"
            tax_note = f"  (total c/Tax: ${d['total_cost']:,.2f}  Tax: ${tax:,.2f})"
        else:
            cost_str = f"${cost_ex_tax:>9,.2f} "
            tax_note = ""

        pdp_str = f"${d['pdp_cost']:>8,.2f}" if d["pdp_cost"] else "        —"
        lines.append(
            f"  {d['date']:<12} {cost_str}  {pdp_str}  {d['max_nodes']:>5}{ggg_marker}{no_payload}  {evt_summary}{tax_note}"
        )
    lines.append("")

    # ── Anomalias recorrentes ──────────────────────────────────────────────────
    if summary["recurring_anomalies"]:
        lines += [
            "─" * 72,
            "  ANOMALIAS RECORRENTES  (apareceram em múltiplos dias de referência)",
            "─" * 72,
        ]
        for r in summary["recurring_anomalies"][:10]:
            days_str = ", ".join(sorted(r["days"]))
            lines.append(f"  [{r['count']}x]  {r['key']}")
            lines.append(f"         Dias      : {days_str}")
            lines.append(f"         Delta acum.: ${r['total_delta_usd']:,.2f}")
            lines.append("")

    # ── Gaps de cobertura ──────────────────────────────────────────────────────
    if summary["days_without_payload"]:
        lines += [
            "─" * 72,
            "  DIAS SEM ANÁLISE DIÁRIA  (payload não gerado — cobertura parcial)",
            "─" * 72,
        ]
        lines.append("  " + ", ".join(sorted(summary["days_without_payload"])))
        lines.append("")

    # ── Narrativa IA ───────────────────────────────────────────────────────────
    if ai_text:
        lines += [
            "─" * 72,
            f"  ANÁLISE IA — {model_id}",
            "─" * 72,
            "",
            ai_text.strip(),
            "",
        ]
    else:
        lines += [
            "─" * 72,
            "  ANÁLISE IA — não habilitada (use --enable-bedrock para gerar narrativa executiva)",
            "─" * 72,
            "",
        ]

    lines += ["=" * 72, "  fim do relatório mensal", "=" * 72]

    path.write_text("\n".join(lines), encoding="utf-8")


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    args = parse_args()
    month = args.month
    export_dir = Path(args.export_dir)
    output_dir = Path(args.output_dir)
    output_root = Path(config.OUTPUT_DIR)

    print(f"[monthly] Mês: {month}")

    # Carregar CSVs mensais
    all_csv = _resolve_csv_path(export_dir, config.MONTHLY_EXPORT_ALL_SERVICES_FILENAME, month)
    pdp_csv = _resolve_csv_path(export_dir, config.MONTHLY_EXPORT_ONLY_PDP_FILENAME, month)

    print(f"[monthly] CSV todos os serviços: {all_csv}")
    all_totals, all_by_day = load_monthly_csv(all_csv)
    if not all_totals:
        print(f"[monthly] ERRO: CSV não encontrado ou vazio: {all_csv}")
        sys.exit(1)

    print(f"[monthly] CSV PDP             : {pdp_csv}")
    pdp_totals, pdp_by_day = load_monthly_csv(pdp_csv)

    # Carregar payloads diários
    print(f"[monthly] Lendo payloads em {output_root}/")
    anomalies_by_day = load_daily_payloads(month, output_root)
    print(f"[monthly] Payloads encontrados: {len(anomalies_by_day)} dias")

    # Carregar calendário de eventos
    print("[monthly] Lendo calendário de eventos...")
    events_by_day = load_events_for_month(month)
    total_events = sum(len(v) for v in events_by_day.values())
    print(f"[monthly] Eventos encontrados : {total_events} em {len(events_by_day)} dias")

    # Consolidar resumo mensal
    summary = build_monthly_summary(
        month, all_totals, all_by_day, pdp_totals, pdp_by_day,
        anomalies_by_day, events_by_day,
    )

    # Bedrock (opcional)
    ai_text: Optional[str] = None
    if args.enable_bedrock:
        print(f"[monthly] Chamando Bedrock ({args.bedrock_model})...")
        try:
            context_op = _load_bedrock_context()
            prompt = _build_monthly_prompt(summary, context_op)
            ai_text = invoke_bedrock(prompt, args.bedrock_model, args.aws_profile, args.bedrock_region)
            print("[monthly] Resposta Bedrock recebida.")
        except Exception as exc:
            print(f"[monthly] AVISO: Bedrock falhou — {exc}")

    # Gerar TXT
    out_path = output_dir / f"relatorio_mensal_{month}.txt"
    write_txt_report(out_path, summary, ai_text, args.bedrock_model)
    print(f"[monthly] Relatório gerado: {out_path}")

    # Notificação SNS (opcional)
    if args.sns_topic_arn and args.s3_bucket:
        try:
            sys.path.insert(0, str(SRC_ROOT))
            from integrations import notifier
            csv_all = export_dir / config.MONTHLY_EXPORT_ALL_SERVICES_FILENAME.replace(".csv", f"-{month}.csv")
            csv_pdp = export_dir / config.MONTHLY_EXPORT_ONLY_PDP_FILENAME.replace(".csv", f"-{month}.csv")
            notifier.notify_monthly_report(
                topic_arn=args.sns_topic_arn,
                bucket=args.s3_bucket,
                s3_prefix=args.s3_prefix,
                csv_all_path=str(csv_all),
                csv_pdp_path=str(csv_pdp),
                aws_profile=args.aws_profile or None,
                aws_region=config.COST_EXPLORER_REGION,
                month=month,
                total_cost=summary["total_cost_usd"],
                pdp_cost=summary["pdp_cost_usd"],
                avg_daily=summary["avg_daily_cost_usd"],
            )
        except Exception as exc:
            print(f"[notifier] AVISO: falha no envio SNS — {exc}")


if __name__ == "__main__":
    main()
