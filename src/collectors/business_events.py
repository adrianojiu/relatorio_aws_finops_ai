"""
Collect business event and push calendar context from the Claro TV+ spreadsheet.
"""

from datetime import date, datetime
import os

from openpyxl import load_workbook

import config


REQUIRED_HEADERS = {
    "Data": "event_date",
    "Horário de disparo": "push_time",
    "Horário de exibição": "display_time",
    "Alcance": "reach",
    "Nodes": "nodes",
    "Destaque": "highlight",
    "Tipo de ação (media)": "action_type",
}

REACH_ORDER = {
    "P": 1,
    "M": 2,
    "G": 3,
    "GG": 4,
    "GGG": 5,
    "4xG": 6,
    "baixo": 1,
}


def load_business_event_calendar(start_date, end_date, reference_day):
    """
    Load business event and push rows for the analysis window.
    """
    calendar_file = config.BUSINESS_EVENT_CALENDAR_FILE
    if not os.path.exists(calendar_file):
        return {
            "enabled": False,
            "source_file": calendar_file,
            "notes": [
                "Arquivo de calendario de eventos/push nao encontrado; sem contexto adicional de negocio.",
            ],
            "window_days": [],
            "reference_day": None,
        }

    start_day = _to_date(start_date)
    end_day = _to_date(end_date)
    reference_day_value = _to_date(reference_day)

    workbook = load_workbook(calendar_file, read_only=True, data_only=True)
    events = []
    seen_keys = set()

    for worksheet in workbook.worksheets:
        header_map = _build_header_map(worksheet)
        if not header_map:
            continue

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            event_day = _to_date(row[header_map["event_date"]])
            action_type = _normalize_action_type(row[header_map["action_type"]])

            if not event_day or action_type not in {"push", "sem push"}:
                continue
            if event_day < start_day or event_day > end_day:
                continue

            push_time = _format_time_value(row[header_map["push_time"]])
            display_time = _format_time_value(row[header_map["display_time"]])
            highlight = _clean_text(row[header_map["highlight"]])
            key = (event_day.isoformat(), action_type, push_time, display_time, highlight)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            events.append(
                {
                    "source_sheet": worksheet.title,
                    "event_date": event_day.isoformat(),
                    "action_type": action_type,
                    "push_time": push_time,
                    "display_time": display_time,
                    "reach": _clean_text(row[header_map["reach"]]),
                    "nodes": _to_int_or_none(row[header_map["nodes"]]),
                    "highlight": highlight,
                }
            )

    window_days = _summarize_days(events)
    reference_summary = next(
        (day for day in window_days if day["date"] == reference_day_value.isoformat()),
        None,
    )

    return {
        "enabled": True,
        "source_file": calendar_file,
        "notes": [
            "Use este calendario apenas como contexto de negocio para eventos, push, aumento de plays e scaling agendado.",
            "Nao use este calendario como evidencia direta de contagem ou custo de AWS End User Messaging/SMS.",
        ],
        "window_days": window_days,
        "reference_day": reference_summary,
    }


def _build_header_map(worksheet):
    first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        return None

    header_map = {}
    for index, value in enumerate(first_row):
        normalized_value = _clean_text(value)
        if normalized_value in REQUIRED_HEADERS:
            header_map[REQUIRED_HEADERS[normalized_value]] = index

    if set(REQUIRED_HEADERS.values()).issubset(header_map):
        return header_map
    return None


def _summarize_days(events):
    grouped = {}
    for event in sorted(events, key=lambda item: (item["event_date"], item["push_time"] or "", item["highlight"] or "")):
        day_summary = grouped.setdefault(
            event["event_date"],
            {
                "date": event["event_date"],
                "event_count": 0,
                "push_count": 0,
                "sem_push_count": 0,
                "max_nodes": None,
                "highest_reach": None,
                "events": [],
            },
        )
        day_summary["event_count"] += 1
        if event["action_type"] == "push":
            day_summary["push_count"] += 1
        elif event["action_type"] == "sem push":
            day_summary["sem_push_count"] += 1

        if event["nodes"] is not None:
            current_max = day_summary["max_nodes"] or 0
            day_summary["max_nodes"] = max(current_max, event["nodes"])

        if _reach_rank(event["reach"]) > _reach_rank(day_summary["highest_reach"]):
            day_summary["highest_reach"] = event["reach"]

        day_summary["events"].append(
            {
                "action_type": event["action_type"],
                "push_time": event["push_time"],
                "display_time": event["display_time"],
                "reach": event["reach"],
                "nodes": event["nodes"],
                "highlight": event["highlight"],
            }
        )

    return list(grouped.values())


def _reach_rank(value):
    if value is None:
        return 0
    return REACH_ORDER.get(str(value).strip(), 0)


def _normalize_action_type(value):
    cleaned = _clean_text(value)
    return cleaned.lower() if cleaned else None


def _format_time_value(value):
    if value is None:
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    cleaned = _clean_text(value)
    return cleaned if cleaned else None


def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _to_int_or_none(value):
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _clean_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None
