"""
Excel report renderer
"""

import openpyxl
from openpyxl.styles import Font
import os
import config
import utils

def write_excel_report(data_inicio, data_fim, ultimo_dia, custo_medio, variacao_media,
                       aumentaram, reduziram, top_costs, daily_costs=None, sms_por_dia=None,
                       total_sms_7d=None, media_sms_7d=None, enriched_anomalies=None):
    """
    Write Excel report
    """
    output_dir = utils.ensure_output_dir()
    output_file = os.path.join(output_dir, utils.get_output_filename("relatorio_custos", "xlsx"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"

    linha = 1
    ws[f"A{linha}"] = f"Relatório de Custos AWS (sem TAX) - {data_inicio} a {data_fim}"
    ws[f"A{linha}"].font = Font(bold=True, size=14)

    linha += 2
    ws[f"A{linha}"] = "Custo médio por dia:"
    ws[f"B{linha}"] = round(custo_medio, 2)

    linha += 1
    ws[f"A{linha}"] = f"Variação média do custo entre os dias {data_inicio} e {data_fim}:"
    ws[f"B{linha}"] = round(variacao_media, 2)

    if daily_costs:
        linha += 2
        ws[f"A{linha}"] = "Custos por dia na janela:"
        ws[f"A{linha}"].font = Font(bold=True)
        linha += 1
        ws.append(["Data", "Custo US$"])
        for date_value, cost in daily_costs:
            ws.append([date_value, round(cost, 2)])

    # Aumentaram
    linha = ws.max_row + 2
    ws[f"A{linha}"] = f"Serviços que aumentaram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
    ws[f"A{linha}"].font = Font(bold=True)
    linha += 1
    ws.append(["Serviço", "Variação US$", "Variação %"])
    for svc, row in aumentaram.iterrows():
        ws.append([svc, round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

    # Reduziram
    linha = ws.max_row + 2
    ws[f"A{linha}"] = f"Serviços que reduziram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
    ws[f"A{linha}"].font = Font(bold=True)
    linha += 1
    ws.append(["Serviço", "Variação US$", "Variação %"])
    for svc, row in reduziram.iterrows():
        ws.append([svc, round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

    # TOP services
    linha = ws.max_row + 2
    ws[f"A{linha}"] = f"TOP {config.TOP_N_SERVICES} serviços em {ultimo_dia} (maiores custos):"
    ws[f"A{linha}"].font = Font(bold=True)
    linha += 1
    rank = 1
    for svc, val in top_costs.items():
        ws.append([f"{rank}. {svc}", round(val, 2)])
        rank += 1

    # SMS if available
    if sms_por_dia is not None:
        linha = ws.max_row + 2
        ws[f"A{linha}"] = f"AWS End User Messaging – Últimos 7 dias:"
        ws[f"A{linha}"].font = Font(bold=True)
        linha += 1
        ws.append(["Data", "Custo US$"])
        for _, row in sms_por_dia.iterrows():
            ws.append([row["Data"], round(row["Custo($)"], 2)])
        ws.append(["Total últimos 7 dias", round(total_sms_7d, 2)])
        ws.append(["Média diária (7 dias)", round(media_sms_7d, 2)])

    if enriched_anomalies:
        linha = ws.max_row + 2
        ws[f"A{linha}"] = "Anomalias correlacionadas para Bedrock"
        ws[f"A{linha}"].font = Font(bold=True)
        linha += 1
        ws.append(["Serviço", "Usage Type", "Custo dia", "Média 7d", "Delta %", "Recurso candidato"])
        for anomaly in enriched_anomalies:
            resources = anomaly.get("resources") or [{}]
            resource = resources[0]
            ws.append([
                anomaly["service"],
                anomaly["usage_type"],
                round(anomaly["cost_today"], 2),
                round(anomaly["avg_7d"], 2),
                f"{anomaly['delta_pct']:.2f}%",
                resource.get("resource_id", "nao identificado"),
            ])

    wb.save(output_file)
    print(f"Relatório Excel salvo em: {output_file}")
    return output_file
