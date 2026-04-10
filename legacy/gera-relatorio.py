import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import os

# ===== Variaveis para execucao =====
csv_file = "csv/02-12-2025.csv"
data_inicio = "2025-11-23"
data_fim = "2025-11-30"
ultimo_dia = "2025-11-30"

# ===== Carregar CSV =====
df = pd.read_csv(csv_file)

# Filtrar período
df_periodo = df[df["Service"].between(data_inicio, data_fim)].copy()

# Calcular custo médio diário
custo_medio = df_periodo["Total costs($)"].mean()

# Valor do último dia
custo_ultimo_dia = df_periodo.loc[df_periodo["Service"] == ultimo_dia, "Total costs($)"].values[0]

# Variação em relação à média
variacao_media = custo_ultimo_dia - custo_medio

# Calcular médias por serviço (ignorar colunas Service e Total costs)
servicos = df_periodo.drop(columns=["Service", "Total costs($)"])
medias = servicos.mean()

# Valores do último dia
ultimo = df.loc[df["Service"] == ultimo_dia, servicos.columns].iloc[0]

# Montar dataframe com variações
df_var = pd.DataFrame({
    "Média": medias,
    "Último dia": ultimo,
})
df_var["Variação US$"] = df_var["Último dia"] - df_var["Média"]
df_var["Variação %"] = (df_var["Variação US$"] / df_var["Média"]) * 100

# Aumentaram e reduziram
aumentaram = df_var[df_var["Variação US$"] > 0].sort_values("Variação US$", ascending=False)
reduziram = df_var[df_var["Variação US$"] < 0].sort_values("Variação US$")

# TOP 5 serviços do último dia
top_costs = ultimo.sort_values(ascending=False).head(5)

# ===== Criar Excel =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Relatório"

linha = 1
ws[f"A{linha}"] = f"Relatório de Custos - {data_inicio} a {data_fim}"
ws[f"A{linha}"].font = Font(bold=True, size=14)

# --- Resumo ---
linha += 2
ws[f"A{linha}"] = "Custo médio por dia:"
ws[f"B{linha}"] = round(custo_medio, 2)

linha += 1
ws[f"A{linha}"] = f"Variação média do custo em relação à média entre os dias {data_inicio} e {data_fim}:"
ws[f"B{linha}"] = round(variacao_media, 2)

# --- Aumentaram ---
linha += 2
ws[f"A{linha}"] = f"Serviços que aumentaram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
ws[f"A{linha}"].font = Font(bold=True)

linha += 1
ws.append(["Serviço", "Variação US$", "Variação %"])
for svc, row in aumentaram.iterrows():
    ws.append([svc.replace("($)", ""), round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

# --- Reduziram ---
linha = ws.max_row + 2
ws[f"A{linha}"] = f"Serviços que reduziram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
ws[f"A{linha}"].font = Font(bold=True)

linha += 1
ws.append(["Serviço", "Variação US$", "Variação %"])
for svc, row in reduziram.iterrows():
    ws.append([svc.replace("($)", ""), round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

# --- TOP 3 ---
linha = ws.max_row + 2
ws[f"A{linha}"] = f"TOP 3 serviços em {ultimo_dia} por custo US$ (maiores custos):"
ws[f"A{linha}"].font = Font(bold=True)

linha += 1
rank = 1
for svc, val in top_costs.items():
    ws.append([f"{rank}. {svc.replace('($)', '')}", round(val, 2)])
    rank += 1

# ===== Salvar Excel =====
data_atual = datetime.now().strftime("%Y-%m-%d")
output_dir = "relatorios"
os.makedirs(output_dir, exist_ok=True)
output_file = os.path.join(output_dir, f"relatorio_custos_{data_atual}.xlsx")

wb.save(output_file)
print(f"Relatório Excel salvo em: {output_file}")

# ===== Gerar TXT =====
output_txt = os.path.join(output_dir, f"relatorio_custos_{data_atual}.txt")

def format_row(svc, var_usd, var_pct):
    return f"{svc.ljust(25)} {str(var_usd).rjust(10)} {str(var_pct).rjust(10)}"

with open(output_txt, "w", encoding="utf-8") as f:
    f.write(f"Custo médio por dia: US$ {custo_medio:,.2f}\n\n")
    f.write(f"Variação média do custo em relação à média entre os dias {data_inicio} e {data_fim}: US$ {variacao_media:,.2f}\n")
    f.write("Tendência de estabilidade diária próxima de 0%.\n\n")

    # --- Aumentaram ---
    f.write(f"Serviços que aumentaram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:\n\n")
    f.write(format_row("Serviço", "Variação US$", "Variação %") + "\n")
    for svc, row in aumentaram.iterrows():
        f.write(format_row(svc.replace("($)",""),
                           f"+{row['Variação US$']:.2f}",
                           f"+{row['Variação %']:.2f}%") + "\n")

    # --- Reduziram ---
    f.write(f"\nServiços que reduziram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:\n\n")
    f.write(format_row("Serviço", "Variação US$", "Variação %") + "\n")
    for svc, row in reduziram.iterrows():
        f.write(format_row(svc.replace("($)",""),
                           f"{row['Variação US$']:.2f}",
                           f"{row['Variação %']:.2f}%") + "\n")

    # --- TOP 3 ---
    f.write(f"\nTOP 5 serviços em {ultimo_dia} por custo US$ (maiores custos):\n\n")
    for i, (svc, val) in enumerate(top_costs.items(), start=1):
        f.write(f"{i}. {svc.replace('($)','')}: US$ {val:,.2f}\n")

print(f"Relatório TXT salvo em: {output_txt}")
