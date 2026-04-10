import boto3
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta
import os

# ============================================================
# CONFIGURAÇÕES GERAIS
# ============================================================
AWS_PROFILE = "prd-ciam"  # altere se usar outro profile local
REGION = "us-east-1"

# Cria sessão boto3 usando o profile local
session = boto3.Session(profile_name=AWS_PROFILE, region_name=REGION)
ce = session.client("ce")

# ============================================================
# PERÍODO DE ANÁLISE (de 9 dias atrás até 2 dias atrás)
# ============================================================
hoje = datetime.now()
data_inicio = (hoje - timedelta(days=9)).strftime("%Y-%m-%d")
data_fim = (hoje - timedelta(days=2)).strftime("%Y-%m-%d")
ultimo_dia = (hoje - timedelta(days=2)).strftime("%Y-%m-%d")

# ============================================================
# CONSULTA AWS COST EXPLORER
# ============================================================
print(f"Coletando custos de {data_inicio} até {data_fim} (sem TAX)...")

response = ce.get_cost_and_usage(
    TimePeriod={
        "Start": data_inicio,
        "End": (hoje - timedelta(days=1)).strftime("%Y-%m-%d")  # AWS exige fim exclusivo
    },
    Granularity="DAILY",
    Metrics=["UnblendedCost"],
    GroupBy=[
        {"Type": "DIMENSION", "Key": "SERVICE"}
    ],
    # Dados serao retornads com base no filtro
    Filter={
        'And': [
            {
                'Not': {
                    'Dimensions': {
                        'Key': 'SERVICE',
                        'Values': ['']  # Servico a excluir entre aspas simples e separar por virgula
                    }
                }
            },
            {
                'Not': {
                    'Dimensions': {
                        'Key': 'RECORD_TYPE',
                        'Values': ['Credit', 'Refund', 'Tax']  # Exclusao de creditos e reembolsos
                    }
                }
            }
        ]
    }
)

# ============================================================
# CONVERSÃO PARA DATAFRAME E FILTRO SEM TAX
# ============================================================
rows = []
for day in response["ResultsByTime"]:
    date = day["TimePeriod"]["Start"]
    for group in day["Groups"]:
        service = group["Keys"][0]
        amount = float(group["Metrics"]["UnblendedCost"]["Amount"])
        rows.append({"Data": date, "Serviço": service, "Custo($)": amount})

df = pd.DataFrame(rows)
if df.empty:
    raise Exception("Nenhum dado retornado do Cost Explorer (após remover TAX).")

# ============================================================
# TABELA PIVOTADA E CÁLCULOS DE MÉDIA E VARIAÇÃO
# ============================================================
df_pivot = df.pivot(index="Data", columns="Serviço", values="Custo($)").fillna(0).reset_index()
df_pivot.rename(columns={"Data": "Service"}, inplace=True)

df_periodo = df_pivot[(df_pivot["Service"] >= data_inicio) & (df_pivot["Service"] <= data_fim)].copy()
total_por_dia = df_periodo.drop(columns=["Service"]).sum(axis=1)
df_periodo["Total costs($)"] = total_por_dia

custo_medio = df_periodo["Total costs($)"].mean()
custo_ultimo_dia = df_periodo.loc[df_periodo["Service"] == ultimo_dia, "Total costs($)"].values[0]
variacao_media = custo_ultimo_dia - custo_medio

servicos = df_periodo.drop(columns=["Service", "Total costs($)"])
medias = servicos.mean()
ultimo = df_periodo.loc[df_periodo["Service"] == ultimo_dia, servicos.columns].iloc[0]

df_var = pd.DataFrame({
    "Média": medias,
    "Último dia": ultimo,
})
df_var["Variação US$"] = df_var["Último dia"] - df_var["Média"]
df_var["Variação %"] = (df_var["Variação US$"] / df_var["Média"]) * 100

aumentaram = df_var[df_var["Variação US$"] > 0].sort_values("Variação US$", ascending=False)
reduziram = df_var[df_var["Variação US$"] < 0].sort_values("Variação US$")
top_costs = ultimo.sort_values(ascending=False).head(5)

# ============================================================
# NOVO: AWS END USER MESSAGING - ÚLTIMOS 7 DIAS
# (Usa o df já coletado; NÃO faz nova chamada ao Cost Explorer)
# ============================================================
servico_sms = "AWS End User Messaging"
inicio_7d = (hoje - timedelta(days=8)).strftime("%Y-%m-%d")
fim_7d = (hoje - timedelta(days=1)).strftime("%Y-%m-%d")

df_sms_7d = df[
    (df["Serviço"] == servico_sms) &
    (df["Data"] >= inicio_7d) &
    (df["Data"] <= fim_7d)
].copy()

sms_por_dia = (
    df_sms_7d
    .groupby("Data")["Custo($)"]
    .sum()
    .reset_index()
    .sort_values("Data")
)

total_sms_7d = sms_por_dia["Custo($)"].sum()
media_sms_7d = sms_por_dia["Custo($)"].mean()

# ============================================================
# CRIAR RELATÓRIO EXCEL
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Relatório"

linha = 1
ws[f"A{linha}"] = f"Relatório de Custos AWS (sem TAX) - {data_inicio} a {data_fim}"
ws[f"A{linha}"].font = Font(bold=True, size=14)

linha += 2
ws[f"A{linha}"] = "Custo médio por dia:"
ws[f"B{linha}"] = round(custo_medio, 2)

linha += 1
ws[f"A{linha}"] = f"Variação média do custo entre os dias {data_inicio} e {data_fim}:"
ws[f"B{linha}"] = round(variacao_media, 2)

# --- Aumentaram ---
linha += 2
ws[f"A{linha}"] = f"Serviços que aumentaram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
ws[f"A{linha}"].font = Font(bold=True)
linha += 1
ws.append(["Serviço", "Variação US$", "Variação %"])
for svc, row in aumentaram.iterrows():
    ws.append([svc, round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

# --- Reduziram ---
linha = ws.max_row + 2
ws[f"A{linha}"] = f"Serviços que reduziram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:"
ws[f"A{linha}"].font = Font(bold=True)
linha += 1
ws.append(["Serviço", "Variação US$", "Variação %"])
for svc, row in reduziram.iterrows():
    ws.append([svc, round(row["Variação US$"], 2), f"{row['Variação %']:.2f}%"])

# --- TOP 3 ---
linha = ws.max_row + 2
ws[f"A{linha}"] = f"TOP 3 serviços em {ultimo_dia} (maiores custos):"
ws[f"A{linha}"].font = Font(bold=True)
linha += 1
rank = 1
for svc, val in top_costs.items():
    ws.append([f"{rank}. {svc}", round(val, 2)])
    rank += 1

# ============================================================
# NOVO: Sessão no FINAL do Excel (abaixo do TOP)
# ============================================================
linha = ws.max_row + 2
ws[f"A{linha}"] = f"AWS End User Messaging – Últimos 7 dias ({inicio_7d} a {ultimo_dia}):"
ws[f"A{linha}"].font = Font(bold=True)

linha += 1
ws.append(["Data", "Custo US$"])

for _, row in sms_por_dia.iterrows():
    ws.append([row["Data"], round(row["Custo($)"], 2)])

ws.append(["Total últimos 7 dias", round(total_sms_7d, 2)])
ws.append(["Média diária (7 dias)", round(media_sms_7d, 2)])

# ============================================================
# SAÍDA DOS RELATÓRIOS
# ============================================================
data_atual = datetime.now().strftime("%Y-%m-%d")
output_dir = "relatorios_costexplorer"
os.makedirs(output_dir, exist_ok=True)

output_file = os.path.join(output_dir, f"relatorio_custos_sem_tax_{data_atual}.xlsx")
wb.save(output_file)
print(f"Relatório Excel salvo em: {output_file}")

# ============================================================
# RELATÓRIO TXT
# ============================================================
output_txt = os.path.join(output_dir, f"relatorio_custos_sem_tax_{data_atual}.txt")

def format_row(svc, var_usd, var_pct):
    return f"{svc.ljust(28)} {str(var_usd).rjust(10)} {str(var_pct).rjust(10)}"

with open(output_txt, "w", encoding="utf-8") as f:
    f.write(f"Custo médio por dia: US$ {custo_medio:,.2f}\n\n")
    f.write(f"Variação média do custo em relação à média entre os dias {data_inicio} e {data_fim}: US$ {variacao_media:,.2f}\n")
    f.write("Tendência de estabilidade diária próxima de 0%.\n\n")

    # --- Aumentaram ---
    f.write(f"Serviços que aumentaram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:\n\n")
    f.write(format_row("Serviço", "Variação US$", "Variação %") + "\n")
    for svc, row in aumentaram.iterrows():
        f.write(format_row(
            svc.replace("($)", ""),
            f"+{row['Variação US$']:.2f}",
            f"+{row['Variação %']:.2f}%"
        ) + "\n")

    # --- Reduziram ---
    f.write(f"\nServiços que reduziram o custo em relação à média entre os dias {data_inicio} e {data_fim} US$:\n\n")
    f.write(format_row("Serviço", "Variação US$", "Variação %") + "\n")
    for svc, row in reduziram.iterrows():
        f.write(format_row(
            svc.replace("($)", ""),
            f"{row['Variação US$']:.2f}",
            f"{row['Variação %']:.2f}%"
        ) + "\n")

    # --- TOP 3 ---
    f.write(f"\nTOP 5 serviços em {ultimo_dia} por custo US$ (maiores custos):\n\n")
    for i, (svc, val) in enumerate(top_costs.items(), start=1):
        f.write(f"{i}. {svc.replace('($)','')}: US$ {val:,.2f}\n")

    # ========================================================
    # NOVO: AWS End User Messaging - Últimos 7 dias (no FINAL)
    # ========================================================
    f.write(f"\n\nAWS End User Messaging – Últimos 7 dias ({inicio_7d} a {ultimo_dia}):\n\n")
    for _, row in sms_por_dia.iterrows():
        f.write(f"{row['Data']}: US$ {row['Custo($)']:,.2f}\n")

    f.write(f"\nTotal últimos 7 dias: US$ {total_sms_7d:,.2f}\n")
    f.write(f"Média diária (7 dias): US$ {media_sms_7d:,.2f}\n")

print(f"Relatório TXT salvo em: {output_txt}")
